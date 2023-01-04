import io
import zipfile
from typing import List

import openpyxl
import pandas as pd
import requests
from google.cloud.bigquery import SchemaField
from prefect import flow, get_run_logger, task
from prefect.task_runners import ConcurrentTaskRunner
from prefect_gcp import GcpCredentials
from prefect_gcp.bigquery import (
    bigquery_create_table,
    bigquery_insert_stream,
    bigquery_query,
)


class Config:
    arbitrary_types_allowed = True


config = {
    "uk": {
        "url_daily": "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/latest-yield-curve-data.zip?la=en&hash=89B8A093FA97EF7DD79382044E15867840E45204",
        "url_historical": "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/glcnominalddata.zip",
        "excel_name": "GLC Nominal daily data current month.xlsx",
        "sheet_name": "4. spot curve",
    },
    "us": {
        "url_historical": "https://home.treasury.gov/system/files/276/yield-curve-rates-1990-2021.csv",
        "url_latest_year": "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/daily-treasury-rates.csv/2022/all?type=daily_treasury_yield_curve&field_tdr_date_value=2022&page&_format=csv",
    },
    "max_records_per_request": 500,
}


def chunker(seq: List, size: int) -> List:
    """
    Split list into chunks

    Args:
        seq (List): List to split
        size (int): Chunk size

    Returns:
        List of chunks

    """
    return [seq[pos : pos + size] for pos in range(0, len(seq), size)]


@task(retries=2, retry_delay_seconds=60)
def request_zip(url: str) -> bytes:
    """
    Use requests to get zip file as bytes

    Args:
        url (str): url to retrieve zip

    Returns:
        bytes

    """
    response = requests.get(url)
    return response.content


@task(retries=2, retry_delay_seconds=60)
def request_csv(url: str) -> pd.DataFrame:
    """
    Use pandas to retrieve csv from url

    Args:
        url (str): url to retrieve csv

    Returns:
        CSV formatted as pd.DataFrame

    """
    curve = pd.read_csv(url)
    return curve


def open_zip(response: bytes) -> zipfile.ZipFile:
    """
    Load zip file bytes into zipfile.ZipFile object

    Args:
        response (bytes): zip file bytes

    Returns:
        zipfile.ZipFile object

    """
    zip = zipfile.ZipFile(io.BytesIO(response), "r")
    return zip


def read_sheet(zip_file: zipfile.ZipFile) -> pd.DataFrame:
    """
    Reads excel sheet into pd.DataFrame

    Args:
        zip_file (zipfile.ZipFile): zipfile.ZipFile object containing multiple excel workbooks

    Returns:
        pd.DataFrame

    """
    excel_file = zip_file.read(config["uk"]["excel_name"])
    curve = pd.read_excel(excel_file, sheet_name=config["uk"]["sheet_name"])
    return curve


@task
def read_historical_sheet(
    zip_file: zipfile.ZipFile, file: str
) -> pd.DataFrame:
    """
    Find spot curve excel sheet and read into pd.DataFrame

    Args:
        zip_file (zipfile.ZipFile): zipfile.ZipFile object containing multiple excel workbooks
        file (str): excel file within zip to read

    Returns:
        pd.DataFrame containing spot curve sheet

    """
    excel_file = zip_file.read(file)
    wb = openpyxl.load_workbook(io.BytesIO(excel_file))
    sheetnames = wb.sheetnames
    # Different years have different sheet names
    sheet = [s for s in sheetnames if "spot curve" in s][0]
    curve = pd.read_excel(excel_file, sheet_name=sheet)
    return curve


@task
def find_gilt_curve(curve: pd.DataFrame) -> pd.DataFrame:
    """
    Find gilt spot curve on excel sheet

    Args:
        curve (pd.DataFrame): Excel sheet containing gilt curve

    Returns:
        pd.DataFrame

    """
    curve.columns = curve.iloc[2].values
    curve = curve.iloc[4:].reset_index(drop=True)
    curve = curve.rename(columns={"years:": "date"})
    curve["date"] = pd.to_datetime(curve["date"]).dt.strftime("%Y-%m-%d")
    curve = curve.set_index("date").transpose().reset_index()
    return curve


@task
def find_treasury_curve(curve: pd.DataFrame) -> pd.DataFrame:
    """
    Find treasury curve within pd.DataFrame

    Args:
        curve (pd.DataFrame): Entire csv loaded as pd.DataFrame

    Returns:
        pd.DataFrame with formatted curve values

    """
    curve = curve.rename(columns={"Date": "date"})
    curve["date"] = pd.to_datetime(curve["date"]).dt.strftime("%Y-%m-%d")
    curve = curve.set_index("date").transpose().reset_index()

    def map_tenors(x: str) -> float:
        value, date_format = x.split()
        if date_format == "Mo":
            value = float(value) / 12.0
        return value

    curve["index"] = curve["index"].apply(map_tenors)
    return curve


@task
def single_curve(spot_curve: pd.DataFrame, date: str, location: str) -> List:
    """
    Get curve points of single date

    Args:
        spot_curve (pd.DataFrame): Spot curve for multiple dates
        date (str): date to filter (%Y-%m-%d format)
        location (str): Location for spot curve

    Returns:
        Curve formatted as list of records (tenor, value, date, location)

    """
    latest_curve = spot_curve[["index", date]]
    latest_curve.columns = ["tenor", "value"]
    latest_curve["location"] = location
    latest_curve["date"] = date
    records = latest_curve.to_dict(orient="records")
    return records


@task
def clean_curve(curve: pd.DataFrame, location: str) -> pd.DataFrame:
    """
    Format multiple curves so it can be inserted into BigQuery

    Args:
        curve (pd.DataFrame): DataFrame containing curves for multiple dates
        location (str): Location of curves

    Returns:
        Formatted curves

    """
    curve = curve.dropna(how="all")
    curve = curve.melt(id_vars="index")
    curve = curve.rename(columns={"index": "tenor"})
    curve = curve.dropna()
    curve["location"] = location
    return curve


@task
def latest_month() -> pd.DataFrame:
    """
    Get UK curve values of the latest month

    Returns:
        pd.DataFrame containing latest months value

    """
    response = request_zip(config["uk"]["url_daily"])
    zip_file = open_zip(response)
    curve = read_sheet(zip_file)
    curve = find_gilt_curve(curve)
    curve = clean_curve(curve, location="UK")
    return curve


def create_table(table: str):
    """
    Create BigQuery table if it does not exist
    """
    schema = [
        SchemaField("date", "DATE"),
        SchemaField("location", "STRING"),
        SchemaField("tenor", "FLOAT64"),
        SchemaField("value", "FLOAT64"),
    ]
    bigquery_create_table(
        dataset="market_data",
        table=table,
        schema=schema,
        gcp_credentials=GcpCredentials(),
    )


def query_exisiting_dates(location: str) -> List:
    """
    Query BigQuery for list of curves already inserted for a given location

    Args:
        location (str): Location to query

    Returns:
        List of dates already inserted

    """
    query = (
        "SELECT DISTINCT date FROM `aurora-361016.market_data.spot_curves`"
        "WHERE location = '%s' " % location
    )
    result = bigquery_query(
        query=query,
        gcp_credentials=GcpCredentials(),
        to_dataframe=True,
    )
    result = list(result.date)
    return result


def latest_curve(curve: pd.DataFrame) -> str:
    """
    Get the date of the latest available curve

    Args:
        curve (pd.DataFrame): DataFrame containing recent curves

    Returns:
        Date formatted as str (%Y-%m-%d)

    """
    latest_dates = list(curve.columns)
    latest_dates.remove("index")
    latest_date = pd.to_datetime(latest_dates).max().strftime("%Y-%m-%d")
    return latest_date


@task(retries=2, retry_delay_seconds=15)
def load_single_curve(
    curve: pd.DataFrame, latest_date: str, location: str, table: str
) -> List:
    """
    Insert single curve into BigQuery

    Args:
        curve (pd.DataFrame): DataFrame of recent curves
        latest_date (str): Curve date to insert
        location (str): Location of curve

    Returns:
        List from prefect bigquery_insert_stream

    """
    logger = get_run_logger()
    logger.info(
        "Loading new curve into big query, curve date: %s", latest_date
    )
    records = single_curve(curve, date=latest_date, location=location)
    result = bigquery_insert_stream(
        dataset="market_data",
        table=table,
        records=records,
        gcp_credentials=GcpCredentials(),
    )
    return result


def no_new_curve(latest_date: str) -> str:
    """
    Handling when there is no curve to insert

    Args:
        latest_date (str): date of latest curve

    Returns:
        String to pass to prefect

    """
    logger = get_run_logger()
    logger.info(
        "No new curve found, curve at %s already loaded",
        latest_date,
    )
    result = "No new curve"
    return result


@task(retries=2, retry_delay_seconds=15)
def insert_chunk(chunk: List, table: str):
    """
    Use prefect_gcp to insert into bigquery

    Args:
        chunk (List): Records to insert into BigQuery

    """
    bigquery_insert_stream(
        dataset="market_data",
        table=table,
        records=chunk,
        gcp_credentials=GcpCredentials(),
    )


@flow(name="Daily UK Government Curve", task_runner=ConcurrentTaskRunner())
def monthly_uk_curve():
    """
    Prefect flow that inserts UK Government Spot curve into BigQuery every day
    """
    response = request_zip(config["uk"]["url_daily"])
    zip_file = open_zip(response)
    curve = read_sheet(zip_file)
    curve = find_gilt_curve(curve)
    loaded_dates = query_exisiting_dates(location="UK")
    latest_date = latest_curve(curve)
    if latest_date not in loaded_dates:
        result = load_single_curve(
            curve, latest_date, location="UK", table="spot_curve"
        )
    else:
        result = no_new_curve(latest_date)
    return result


@flow(name="Daily US Treasury Curve", task_runner=ConcurrentTaskRunner())
def monthly_us_curve():
    """
    Prefect flow that inserts US Treasury Par curve into BigQuery every day
    """
    curve = request_csv(config["us"]["url_latest_year"])
    curve = find_treasury_curve(curve)
    loaded_dates = query_exisiting_dates(location="US")
    latest_date = latest_curve(curve)
    if latest_date not in loaded_dates:
        result = load_single_curve(
            curve, latest_date, location="US", table="par_curve"
        )
    else:
        result = no_new_curve(latest_date)
    return result


@flow(name="Subflow backfill single UK excel sheet", validate_parameters=False)
def backfill_uk_subflow(zip_file: zipfile.ZipFile, file: str):
    """
    Subflow to read and insert batch of UK spot curves

    Args:
        zip_file (zipfile.ZipFile): Zip containg all excel files
        file (str): file name to read

    """
    logger = get_run_logger()
    logger.info("Getting curve data from file: %s", file)
    curve = read_historical_sheet(zip_file, file)
    curve = find_gilt_curve(curve)
    curve = clean_curve(curve, location="UK")
    records = curve.to_dict(orient="records")
    for chunk in chunker(records, config["max_records_per_request"]):
        insert_chunk(chunk, table="spot_curve").submit()


@flow(name="Backfill UK Government Curve", task_runner=ConcurrentTaskRunner())
def backfill_uk_curve():
    """
    Prefect flow that backfills entire history of UK Government spot curves into BigQuery
    """
    response = request_zip(config["uk"]["url_historical"])
    zip_file = open_zip(response)
    create_table(table="spot_curve")
    for file in zip_file.namelist():
        backfill_uk_subflow(zip_file, file)
    # UK Historical records only go up to latest month I think
    curve = latest_month()
    records = curve.to_dict(orient="records")
    for chunk in chunker(records, config["max_records_per_request"]):
        insert_chunk(chunk, table="spot_curve").submit()


@flow(name="Backfill US Treasury Curve", task_runner=ConcurrentTaskRunner())
def backfill_us_curve():
    """
    Prefect flow that backfills entire history of US Treasury par curves into BigQuery
    """
    logger = get_run_logger()
    create_table(table="par_curve")
    us_list = ["url_historical", "url_latest_year"]
    for file in us_list:
        logger.info("Getting curve data from file: %s", file)
        curve = request_csv(config["us"][file])
        curve = find_treasury_curve(curve)
        curve = clean_curve(curve, location="US")
        records = curve.to_dict(orient="records")
        for chunk in chunker(records, config["max_records_per_request"]):
            insert_chunk(chunk, table="par_curve").submit()
